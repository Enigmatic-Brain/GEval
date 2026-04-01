"""
run_tests.py
Reads test_cases.xlsx → generates model responses → evaluates completeness
→ writes results.xlsx

Usage:
  export OPENAI_API_KEY=sk-...
  python run_tests.py
"""

import asyncio
import json
import logging
import os
from datetime import datetime

import aiohttp
import openpyxl
from completeness_evaluator_openai import evaluate_completeness
from groundedness_evaluator_openai import evaluate_groundedness

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

OPENAI_URL = "https://api.openai.com/v1/chat/completions"
MODEL      = "gpt-4o"

# ---------------------------------------------------------------------------
# Prompt to auto-generate a model response for a given scenario
# ---------------------------------------------------------------------------

RESPONSE_PROMPT = """\
You are generating a test response to the question below, using only the provided document.
The response quality MUST match the scenario exactly.

Scenario definitions:

  complete
    • Cover EVERY critical fact, figure, condition, threshold, and edge case in the document.
    • Include all supporting details (rates, fees, dates, exceptions, procedures).
    • Use specific numbers and terms from the document — do not paraphrase vaguely.
    • A reader must be able to answer any operational question from your response alone.

  incomplete
    • Write 2–3 short, generic sentences that barely address the question.
    • Do NOT include any specific numbers, thresholds, fees, or named conditions.
    • The response must be so vague it is almost useless for operational purposes.

Scenario: {scenario}
Question: {question}
Document:
{document}

Return only the response text. No preamble, no labels, no explanation.
"""

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def post_openai(session, api_key, prompt):
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model":       MODEL,
        "max_tokens":  1000,
        "temperature": 0,
        "messages":    [{"role": "user", "content": prompt}],
    }
    async with session.post(OPENAI_URL, json=payload, headers=headers,
                            timeout=aiohttp.ClientTimeout(total=60)) as resp:
        data = await resp.json()
    return data["choices"][0]["message"]["content"].strip()


def read_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip().lower().replace(" ", "_") if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(row)]


def write_excel(results, path):
    wb = openpyxl.Workbook()

    # Sheet 1 — summary
    ws1 = wb.active
    ws1.title = "Results"
    headers = [
        "test_id", "document_type", "scenario_type", "expected_label",
        "actual_label", "score", "pass_fail", "total_claims",
        "critical_claims", "supporting_claims",
        "missing_critical", "missing_supporting",
        "question", "model_response", "timestamp",
    ]
    ws1.append(headers)
    for r in results:
        ws1.append([r.get(h, "") for h in headers])

    # Sheet 2 — per-claim breakdown
    ws2 = wb.create_sheet("Claim Breakdown")
    ws2.append(["test_id", "document_type", "scenario_type",
                "claim_id", "coverage_score", "coverage_label", "reason"])
    for r in results:
        for c in r.get("claims", []):
            ws2.append([
                r["test_id"], r["document_type"], r["scenario_type"],
                c["claim_id"], c["coverage_score"], c["coverage_label"], c["reason"],
            ])

    # Basic column widths
    for ws in (ws1, ws2):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    wb.save(path)
    log.info(f"Saved → {path}")


def write_groundedness_excel(results, path):
    wb = openpyxl.Workbook()

    # Sheet 1 — summary
    ws1 = wb.active
    ws1.title = "Groundedness Results"
    headers = [
        "test_id", "document_type", "scenario_type",
        "groundedness_score", "groundedness_label",
        "total_claims", "supported_claims", "not_supported_claims",
        "unsupported_claims",
        "question", "model_response", "document", "timestamp",
    ]
    ws1.append(headers)
    for r in results:
        ws1.append([r.get(h, "") for h in headers])

    # Sheet 2 — per-claim verification breakdown (for debugging)
    ws2 = wb.create_sheet("Claim Verification")
    ws2.append([
        "test_id", "document_type", "claim_id", "claim_text",
        "verification_status", "reason",
    ])
    for r in results:
        for c in r.get("groundedness_details", []):
            ws2.append([
                r["test_id"], r["document_type"],
                c["claim_id"], c["claim_text"],
                c["verification_status"], c["reason"],
            ])

    # Basic column widths
    for ws in (ws1, ws2):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

    wb.save(path)
    log.info(f"Saved → {path}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def main():
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise EnvironmentError("Set OPENAI_API_KEY first.")

    test_cases = read_excel("test_cases.xlsx")
    log.info(f"Loaded {len(test_cases)} test cases")

    results = []
    groundedness_results = []
    async with aiohttp.ClientSession() as session:
        for i, tc in enumerate(test_cases, 1):
            log.info(f"\n[{i}/{len(test_cases)}] {tc['test_id']} | {tc.get('document_type','') } | {tc.get('scenario_type','')}")
            try:
                prompt = ""
                # Step 0: use existing model_response if provided, otherwise generate one
                existing = tc.get("model_response")
                if existing and str(existing).strip():
                    model_response = str(existing).strip()
                    log.info("  Step 0: using model_response from input.")
                else:
                    prompt = RESPONSE_PROMPT.format(
                        scenario=tc.get("scenario_type", "complete"),
                        question=tc["question"],
                        document=tc["document"],
                    )
                    model_response = await post_openai(session, api_key, prompt)
                    log.info("  Step 0: model_response generated.")

                # Steps 1–3: evaluate completeness
                eval_result = await evaluate_completeness(
                    question=tc["question"],
                    model_response=model_response,
                    document=tc["document"],
                    api_key=api_key,
                )

                for c in eval_result["coverage_details"]:
                    log.info(f"  Claim {c['claim_id']} [{c['coverage_label']}] {c['reason']}")

                actual   = eval_result["completeness_label"]
                expected = tc.get("expected_label", "")
                results.append({
                    "test_id":           tc["test_id"],
                    "document_type":     tc.get("document_type", ""),
                    "scenario_type":     tc.get("scenario_type", ""),
                    "expected_label":    expected,
                    "actual_label":      actual,
                    "score":             eval_result["weighted_completeness_score"],
                    "pass_fail":         "PASS" if actual == expected else "FAIL",
                    "total_claims":      eval_result["total_claims"],
                    "critical_claims":   eval_result["critical_claims"],
                    "supporting_claims": eval_result["supporting_claims"],
                    "missing_critical":  "; ".join(eval_result["missing_critical_claims"]),
                    "missing_supporting":"; ".join(eval_result["missing_supporting_claims"]),
                    "question":          tc["question"],
                    "model_response":    model_response,
                    "timestamp":         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "claims":            eval_result["coverage_details"],
                })
                log.info(f"  Completeness={eval_result['weighted_completeness_score']:.2f} | {actual} | {'PASS' if actual == expected else 'FAIL'}")

                # Groundedness evaluation
                ground_result = await evaluate_groundedness(
                    model_response=model_response,
                    document=tc["document"],
                    api_key=api_key,
                )
                groundedness_results.append({
                    "test_id":              tc["test_id"],
                    "document_type":        tc.get("document_type", ""),
                    "scenario_type":        tc.get("scenario_type", ""),
                    "groundedness_score":   ground_result["groundedness_score"],
                    "groundedness_label":   ground_result["groundedness_label"],
                    "total_claims":         ground_result["total_claims"],
                    "supported_claims":     ground_result["supported_claims"],
                    "not_supported_claims": ground_result["not_supported_claims"],
                    "unsupported_claims":   "; ".join(ground_result["unsupported_claim_texts"]),
                    "question":             tc["question"],
                    "model_response":       model_response,
                    "document":             tc["document"],
                    "timestamp":            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "groundedness_details": ground_result["verification_details"],
                })
                log.info(f"  Groundedness={ground_result['groundedness_score']:.2f} | {ground_result['groundedness_label']}")

            except Exception as e:
                log.error(f"  Failed: {e}")
                results.append({
                    "test_id":       tc["test_id"],
                    "document_type": tc.get("document_type", ""),
                    "scenario_type": tc.get("scenario_type", ""),
                    "expected_label":tc.get("expected_label", ""),
                    "actual_label":  f"ERROR: {e}",
                    "pass_fail":     "ERROR",
                    "timestamp":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "claims":        [],
                })
                groundedness_results.append({
                    "test_id":              tc["test_id"],
                    "document_type":        tc.get("document_type", ""),
                    "scenario_type":        tc.get("scenario_type", ""),
                    "groundedness_score":   0.0,
                    "groundedness_label":   f"ERROR: {e}",
                    "total_claims":         0,
                    "supported_claims":     0,
                    "not_supported_claims": 0,
                    "unsupported_claims":   "",
                    "question":             tc.get("question", ""),
                    "model_response":       "",
                    "document":             tc.get("document", ""),
                    "timestamp":            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "groundedness_details": [],
                })

    write_excel(results, "results.xlsx")
    write_groundedness_excel(groundedness_results, "groundedness_results.xlsx")

    # Print quick summary
    total  = len(results)
    passed = sum(1 for r in results if r.get("pass_fail") == "PASS")
    grounded = sum(1 for r in groundedness_results if r.get("groundedness_label") == "Grounded")
    print(f"\nDone: {passed}/{total} completeness passed | {grounded}/{total} grounded")


if __name__ == "__main__":
    asyncio.run(main())
